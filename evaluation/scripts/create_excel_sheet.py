"""
create_excel_sheet.py — Create Annotator Excel Workbook

Builds a two-sheet Excel file ready for human annotation:
  - Sheet 1 "Instructions": how-to guide + topic definitions
  - Sheet 2 "Annotation":   one row per article, with:
      • Checkmark (✓) dropdown for each of the 12 social topics
        — blank = does not apply, ✓ = applies (turns green)
      • "Other" checkmark column + "Other Description" free-text cell
      • "Notes" free-text cell
      • Frozen header row + frozen first 6 info columns
      • Auto-filter on all columns
      • Pipeline predictions pre-filled (✓ where pipeline said Yes)

Usage:
    python3.11 evaluation/scripts/create_excel_sheet.py
    python3.11 evaluation/scripts/create_excel_sheet.py \\
        --input  evaluation/annotation/annotation_sample_1000.csv \\
        --output evaluation/annotation/annotation_sheet.xlsx
"""

import argparse
import os
import sys

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from analysis.topics import TOPICS

TOPIC_IDS = [t["id"] for t in TOPICS]

CHECK = "✓"       # value written when a topic applies
EMPTY = ""        # value when topic does not apply


# ── Shared style helpers ──────────────────────────────────────────────────────

def _thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color.lstrip("#"))


# ── Sheet 1: Instructions ─────────────────────────────────────────────────────

def _build_instructions(ws):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 75
    ws.sheet_view.showGridLines = False

    navy = "1F3864"
    row = 1

    # Title banner
    title = ws.cell(row, 1, "ANNOTATION INSTRUCTIONS — BLACK MEDIA TOPIC STUDY")
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = _fill(navy)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{row}:B{row}")
    ws.row_dimensions[row].height = 35
    row += 2

    # How-to steps
    section = ws.cell(row, 1, "HOW TO ANNOTATE")
    section.font = Font(bold=True, size=12, color=navy)
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    steps = [
        ("1. Open the sheet",
         "Click the 'Annotation' tab at the bottom of this workbook."),
        ("2. Read each article",
         "Review the Title (column D) and Description (column E) for each row. "
         "Click the link in column F to open the full article if needed."),
        ("3. Mark the topics",
         "Click any topic cell (columns G onward). A small dropdown arrow appears — "
         "click it and select ✓ to mark the topic. The cell turns green. "
         "Leave blank if the topic does NOT apply. "
         "Pipeline predictions are pre-filled — remove ✓ if the pipeline was wrong."),
        ("4. Use 'Other' if needed",
         "If the article covers a racial-justice or social-equity topic NOT in our "
         "12 categories, click ✓ in the 'Other' column, then type the topic name "
         "in the 'Other Description' column next to it."),
        ("5. Add notes",
         "Use the 'Notes' column for any uncertainty, ambiguity, or comments."),
        ("6. Save when done",
         "Save the file (Ctrl+S / ⌘+S) and return it to the research team."),
        ("Tip",
         "An article can cover multiple topics — mark ✓ in all that apply. "
         "You only need to mark topics that DO apply; blank means 'No'."),
    ]

    for label, text in steps:
        lbl_cell = ws.cell(row, 1, label)
        lbl_cell.font = Font(bold=True, size=11)
        lbl_cell.alignment = Alignment(vertical="top")

        txt_cell = ws.cell(row, 2, text)
        txt_cell.font = Font(size=11)
        txt_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 36
        row += 1

    row += 1

    # Topic definitions
    section = ws.cell(row, 1, "TOPIC DEFINITIONS")
    section.font = Font(bold=True, size=12, color=navy)
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    for topic in TOPICS:
        hex_col = topic["color"].lstrip("#")
        name_cell = ws.cell(row, 1, topic["name"])
        name_cell.font = Font(bold=True, size=11, color="FFFFFF")
        name_cell.fill = _fill(hex_col)
        name_cell.alignment = Alignment(wrap_text=True, vertical="top")

        desc_cell = ws.cell(row, 2, topic["description"])
        desc_cell.font = Font(size=11)
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 38
        row += 1

    row += 1

    # "Other" explanation
    ws.cell(row, 1, "ABOUT 'OTHER'").font = Font(bold=True, size=12, color=navy)
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    ws.cell(row, 1, "Other ✓").font = Font(bold=True, size=11)
    ws.cell(row, 2,
        "Mark ✓ in the 'Other' column when an article covers a racial justice "
        "or social equity topic that does NOT fit any of the 12 categories above. "
        "Then describe the topic in plain language in the adjacent 'Other Description' cell. "
        "Examples: 'Black mental health', 'LGBTQ+ rights in Black communities', 'Colorism'."
    ).font = Font(size=11)
    ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 55


# ── Sheet 2: Annotation ───────────────────────────────────────────────────────

INFO_HEADERS = ["ID", "Source", "Date", "Title", "Description", "Link"]
INFO_WIDTHS  = [6,    20,       12,     42,       58,             11]

CHECK_FILL = _fill("C6EFCE")   # green
CHECK_FONT = Font(bold=True, size=14, color="375623")
EMPTY_FILL = _fill("FAFAFA")   # near-white
EMPTY_FONT = Font(size=10, color="CCCCCC")
BORDER     = _thin_border()


def _build_annotation(ws, df: pd.DataFrame):
    ws.sheet_view.showGridLines = False

    n_info      = len(INFO_HEADERS)         # 6
    n_topics    = len(TOPIC_IDS)            # 12
    topic_start = n_info + 1               # col 7 = G
    other_col   = topic_start + n_topics   # col 19
    other_desc  = other_col + 1            # col 20
    notes_col   = other_col + 2            # col 21
    last_data   = len(df) + 1

    # ── Column widths ─────────────────────────────────────────────────
    for i, w in enumerate(INFO_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(n_topics):
        ws.column_dimensions[get_column_letter(topic_start + i)].width = 9
    ws.column_dimensions[get_column_letter(other_col)].width  = 9
    ws.column_dimensions[get_column_letter(other_desc)].width = 28
    ws.column_dimensions[get_column_letter(notes_col)].width  = 28

    # ── Header row (row 1) ────────────────────────────────────────────
    navy_fill  = _fill("1F3864")
    navy_font  = Font(bold=True, size=10, color="FFFFFF")
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_i, label in enumerate(INFO_HEADERS, start=1):
        c = ws.cell(1, col_i, label)
        c.font, c.fill, c.alignment, c.border = navy_font, navy_fill, center, BORDER

    for i, tid in enumerate(TOPIC_IDS):
        topic = next(t for t in TOPICS if t["id"] == tid)
        col   = topic_start + i
        # Short 2-line label: number + abbreviated name
        short = topic["name"].replace(" & ", "/").replace(" and ", "/")
        words = short.split()
        # Wrap at ~2 words per line for narrow columns
        label = "\n".join([" ".join(words[j:j+2]) for j in range(0, len(words), 2)])
        c = ws.cell(1, col, label)
        c.font      = Font(bold=True, size=8, color="FFFFFF")
        c.fill      = _fill(topic["color"])
        c.alignment = center
        c.border    = BORDER

    grey_hdr = Font(bold=True, size=9, color="FFFFFF")
    for col, label in [
        (other_col,  "Other"),
        (other_desc, "Other\nDescription"),
        (notes_col,  "Notes"),
    ]:
        c = ws.cell(1, col, label)
        c.font      = grey_hdr
        c.fill      = _fill("595959")
        c.alignment = center
        c.border    = BORDER

    ws.row_dimensions[1].height = 52

    # ── Subheader row (row 2) — full topic names as tooltip-style reference ──
    for i, tid in enumerate(TOPIC_IDS):
        topic = next(t for t in TOPICS if t["id"] == tid)
        col   = topic_start + i
        c = ws.cell(2, col, topic["name"])
        c.font      = Font(italic=True, size=7, color="555555")
        c.fill      = _fill("F5F5F5")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER

    for col_i in range(1, n_info + 1):
        c = ws.cell(2, col_i, "")
        c.fill   = _fill("F0F0F0")
        c.border = BORDER

    for col in [other_col, other_desc, notes_col]:
        c = ws.cell(2, col, "")
        c.fill   = _fill("F5F5F5")
        c.border = BORDER

    ws.row_dimensions[2].height = 22

    # ── Data rows (start at row 3) ────────────────────────────────────
    alt_fill = _fill("EEF2F7")

    for row_i, (_, row) in enumerate(df.iterrows(), start=3):
        row_fill = alt_fill if row_i % 2 == 0 else _fill("FFFFFF")

        # Info columns
        values = [
            row.get("article_id", row_i - 2),
            row.get("source",      ""),
            str(row.get("date",   ""))[:10],
            row.get("title",       ""),
            row.get("description", ""),
            row.get("link",        ""),
        ]
        for col_i, val in enumerate(values, start=1):
            c = ws.cell(row_i, col_i, val)
            c.font      = Font(size=9)
            c.fill      = row_fill
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border    = BORDER

        # Topic checkmark columns
        for j, tid in enumerate(TOPIC_IDS):
            col = topic_start + j
            raw = str(row.get(f"label_{tid}", ""))
            # Convert legacy Yes/No to ✓/blank
            val = CHECK if raw in ("Yes", CHECK, "1", "True") else EMPTY
            c = ws.cell(row_i, col, val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = BORDER
            if val == CHECK:
                c.fill, c.font = CHECK_FILL, CHECK_FONT
            else:
                c.fill, c.font = EMPTY_FILL, EMPTY_FONT

        # Other checkmark column
        other_raw = str(row.get("other", ""))
        other_val = CHECK if other_raw in ("Yes", CHECK, "1") else EMPTY
        c = ws.cell(row_i, other_col, other_val)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BORDER
        if other_val == CHECK:
            c.fill, c.font = CHECK_FILL, CHECK_FONT
        else:
            c.fill, c.font = EMPTY_FILL, EMPTY_FONT

        # Other description + Notes
        for col, key in [(other_desc, "other_category"), (notes_col, "annotator_notes")]:
            c = ws.cell(row_i, col, row.get(key, ""))
            c.font      = Font(size=9)
            c.fill      = row_fill
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border    = BORDER

        ws.row_dimensions[row_i].height = 55

    # ── Dropdown validation (✓ or blank) ─────────────────────────────
    # showDropDown=False means the arrow IS visible (Excel naming quirk)
    dv_topics = DataValidation(
        type="list", formula1=f'"{CHECK},"',
        allow_blank=True, showDropDown=False,
        showErrorMessage=True,
        error="Click the dropdown arrow and select ✓, or leave blank.",
        errorTitle="Invalid entry",
    )
    dv_topics.sqref = (
        f"{get_column_letter(topic_start)}3"
        f":{get_column_letter(topic_start + n_topics - 1)}{last_data}"
    )
    ws.add_data_validation(dv_topics)

    dv_other = DataValidation(
        type="list", formula1=f'"{CHECK},"',
        allow_blank=True, showDropDown=False,
        showErrorMessage=True,
        error="Click the dropdown arrow and select ✓, or leave blank.",
        errorTitle="Invalid entry",
    )
    dv_other.sqref = f"{get_column_letter(other_col)}3:{get_column_letter(other_col)}{last_data}"
    ws.add_data_validation(dv_other)

    # ── Conditional formatting ────────────────────────────────────────
    cf_range = (
        f"{get_column_letter(topic_start)}3"
        f":{get_column_letter(other_col)}{last_data}"
    )
    ws.conditional_formatting.add(cf_range, CellIsRule(
        operator="equal", formula=[f'"{CHECK}"'],
        fill=CHECK_FILL, font=CHECK_FONT
    ))

    # ── Freeze panes (below 2-row header, right of 6 info cols) ──────
    ws.freeze_panes = "G3"

    # ── Auto-filter (row 1 only) ──────────────────────────────────────
    last_col = get_column_letter(notes_col)
    ws.auto_filter.ref = f"A1:{last_col}{last_data}"


# ── Entry point ───────────────────────────────────────────────────────────────

def create_workbook(input_path: str, output_path: str):
    df = pd.read_csv(input_path)
    print(f"\n[Excel] {len(df)} articles loaded from {input_path}")

    wb = openpyxl.Workbook()
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    _build_instructions(ws_inst)

    ws_ann = wb.create_sheet("Annotation")
    _build_annotation(ws_ann, df)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    print(f"\n[Excel] ✓ Workbook saved → {output_path}")
    print(f"        {len(df)} articles | {len(TOPIC_IDS)} topic columns | Other + Other Description + Notes")
    print(f"        Marking: click cell → select ✓ (turns green) | blank = does not apply")
    print(f"\n  Open annotation_sheet.xlsx in Excel or Google Sheets.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Create annotator Excel workbook")
    parser.add_argument("--input",  default="evaluation/annotation/annotation_sample_1000.csv")
    parser.add_argument("--output", default="evaluation/annotation/annotation_sheet.xlsx")
    args = parser.parse_args()
    create_workbook(input_path=args.input, output_path=args.output)
